from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json


#load excel file
wb = load_workbook(filename='output.xlsx')
ws = wb.active

#lege list aamaken
my_list = []


#get last kolom en rij

last_column = len(list(ws.columns))
last_row = len(list(ws.rows))

#loop elke row en kolom

for row in range(1, last_row + 1):
my_dict = {}
for column in range(1, last_column + 1):
column_letter = get_column_letter(column)
if row > 1:
my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value 

#toevoegen dict in de list
my_list.append(my_dict)


# lijst formatteren en omzetten naar json
data = json.dumps(my_list, sort_keys=True, indent=4)
with open('data.json', 'w', encoding='utf-8') as f:
f.write(data)